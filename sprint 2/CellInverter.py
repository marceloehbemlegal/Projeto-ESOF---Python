#! python3
# CellInverter.py - Takes a sheet and shifts the rows into columns and vice versa


import openpyxl


wb = openpyxl.load_workbook('blank_rows.xlsx')
sheet = wb.get_active_sheet()

wbInverted = openpyxl.Workbook()
sheetInverted = wbInverted.get_active_sheet()

sheetList = []

for row in range(1, sheet.max_row+1):
    rowList = [] 
    for column in range(1, sheet.max_column+1): # saves all values from the current row into 'rowList'
        rowList.append(sheet.cell(row=row, column=column).value)
    sheetList.append(rowList) # each element in 'sheetList' will be a list that contains a row from the sheet

for row in range(1, sheet.max_column+1): # copies the original row and prints it as a column in the new sheet
    for column in range(1, sheet.max_row+1):
        sheetInverted.cell(row=row, column=column).value = sheetList[column-1][row-1]

wbInverted.save('inverted_cells.xlsx')
