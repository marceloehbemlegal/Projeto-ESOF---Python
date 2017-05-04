#! python3
# BlankRowInserter.py - Inserts a given amount of rows from the given starting row in a sheet


import openpyxl


wb = openpyxl.load_workbook('multiplication_table.xlsx')
wbWithBlankRow = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheetBlankRow = wbWithBlankRow.get_active_sheet()

N = int(input())
M = int(input())

rowBlank = 1

for row in range(1, wb.max_row+1+M):  # from row 1 to the max row of the original sheet plus the amount to be inserted
    if row == N: # when the current row is the given starting point, skips the given amount of rows
        rowBlank += M
    for column in range(1, wb.max_column+1): # copies the content from the original sheet to the new one
        sheetBlankRow.cell(row=rowBlank, column=column).value = sheet.cell(row=row, column=column).value
    rowBlank += 1

wbWithBlankRow.save('blank_rows.xlsx')
