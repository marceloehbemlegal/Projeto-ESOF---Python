#! python3
# SpreadsheetToText.py - Loads a sheet and copies each column into a separate text file


import openpyxl


wb = openpyxl.load_workbook('multiplication_table.xlsx')
sheet = wb.get_active_sheet()

for column in range(1, sheet.max_column+1):
    file = open('spreadsheetToText_%03d.txt' % column, 'w')  # new file for each new column
    for row in range(1, sheet.max_row+1):
        if sheet.cell(row=row, column=column).value == None: # if cell is blank, prints end of line
            file.write('\n')
        else:
            file.write(str(sheet.cell(row=row, column=column).value)+'\n')
    file.close()
