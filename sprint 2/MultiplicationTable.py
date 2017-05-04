#! python3
# MultiplicationTable.py - Takes a given number and creates a sheet with a multiplication table up to that number


import openpyxl


number = int(input())

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

sheet.freeze_panes = 'B2' # freezes the first row and column

for rowNum in range(1, number+2):
    for colNum in range(1, number+2):
        if not rowNum == colNum == 1:   # leaves first row and column blank
            sheet.cell(row=rowNum, column=colNum).value = (rowNum-1)*(colNum-1)
        if rowNum == 1 and colNum != 1: # fills first row with 1, 2, 3, 4 ...
            sheet.cell(row=rowNum, column=colNum).value = colNum-1
        if colNum == 1 and rowNum != 1: # fills first column with 1//2//3//4 ...
            sheet.cell(row=rowNum, column=colNum).value = rowNum-1

wb.save('multiplication_table.xlsx')
